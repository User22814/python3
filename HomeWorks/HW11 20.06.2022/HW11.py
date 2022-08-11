import openpyxl

workbook = openpyxl.load_workbook("data.xlsx")
worksheet = workbook.active

max_column = worksheet.max_column + 1

set1 = set()
set2 = set()
set3 = set()
# print(type(set1))

for i in range(1, max_column):
    value1 = worksheet.cell(row=2, column=i).value
    value2 = worksheet.cell(row=4, column=i).value
    value3 = worksheet.cell(row=6, column=i).value
    # print(f"value row 2:{value1}")
    # print(f"value row 4:{value2}")
    # print(f"value row 6:{value3}")
    set1.add(value1)
    set2.add(value2)
    set3.add(value3)

# print(f"set1: {set1}")
# print(f"set2: {set2}")
# print(f"set3: {set3}")

# set_intersection_1 = set2.intersection(set1)
# set_intersection_3 = set2.intersection(set3)
set_intersection_1_and_3 = set2.intersection(set1, set3)

# print(set_intersection_1)
# print(set_intersection_3)
print(set_intersection_1_and_3)

# set_difference_1 = set2.difference(set1)
# set_difference_3 = set2.difference(set3)
set_difference_1_and_3 = set2.difference(set1, set3)

# print(set_difference_1)
# print(set_difference_3)
print(set_difference_1_and_3)

set4 = set3

for i in set1:
    set4.add(i)

print(set4)
