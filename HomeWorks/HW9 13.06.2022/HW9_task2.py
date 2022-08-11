import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime


class MyClass:
    def __init__(self, name: str, index: int, value):
        self.name = name
        self.index = index
        self.value = value

    def get_full_info(self):
        print(f"NAME: {self.name}; INDEX: {self.index}; VALUE: {self.value}")


class MyClassForExcel:
    def __init__(self, value1, value2):
        self.col1 = value1
        self.col2 = value2

    def get_full_info(self):
        print(f"1st Column: {self.col1}; 2nd Column: {self.col2}")


array = []

for i in range(1, 1000):
    tmp = MyClass(f'_{i}', i, f'_A{i}')
    array.append(tmp)
print(array)

for i in range(1, 1000):
    array[i - 1].get_full_info()


workbook = openpyxl.load_workbook("List1.xlsx")
worksheet = workbook.active

max_row = worksheet.max_row + 1

array_for_excel = []

for i in range(1, max_row):
    value1 = worksheet.cell(row=i, column=1).value
    value2 = worksheet.cell(row=i, column=2).value
    tmp_excel = MyClassForExcel(value1, value2)
    array_for_excel.append(tmp_excel)

print("\nData from Excel:")
for i in range(1, max_row):
    array_for_excel[i - 1].get_full_info()





