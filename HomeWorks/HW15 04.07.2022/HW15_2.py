import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter

# Read all data from SQL
conn = psycopg2.connect("dbname=new1_postgres_db user=postgres password=qwerty228")
cur = conn.cursor()
cur.execute("""
SELECT * FROM public.new1_postgres_tb
WHERE user_id = '1'
ORDER BY id ASC
""")
records = cur.fetchall()
conn.close()

# Write all data to .xlsx
book = openpyxl.Workbook()
sheet = book.active

titles = ["userId", "id", "title", "body"]
index_col = 1
for title in titles:
    sheet[f"{get_column_letter(index_col)}1"] = str(title)
    index_col += 1

row_index = 2
for row in records:
    col_index = 1
    for column in row:
        sheet[f"{get_column_letter(col_index)}{row_index}"] = str(column)
        col_index += 1
    row_index += 1
book.save('new_table.xlsx')
